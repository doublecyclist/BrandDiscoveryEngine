"""
Brand Discovery Engine
Run:
    pip install -r requirements.txt
    python brand_engine.py --input candidates.csv --output brand_results.xlsx

Optional:
    set YOUTUBE_API_KEY=your_key
    set SERPAPI_KEY=your_key

This tool automates what can be checked safely:
- Domain DNS existence checks
- Search URLs for registrars and social platforms
- Apple Podcasts/iTunes Search API
- Optional YouTube Data API search
- Optional SerpAPI Google search
- Initial scoring model

It does NOT scrape Instagram/Facebook/TikTok/LinkedIn/X directly.
Those results should be reviewed manually using the generated links.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import socket
import sys
import time
import urllib.parse
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import requests
except ImportError:
    requests = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


TLDs = ["com", "co", "net", "org", "guide", "travel", "media", "ai"]
MANUAL_STATUSES = ["Unknown", "Available", "Taken", "Inactive", "Possible conflict", "High conflict", "Premium"]
RISK_STATUSES = ["Unknown", "Low", "Medium", "High"]
RECOMMENDATIONS = ["Research", "Strong candidate", "Possible", "Caution", "Eliminate", "Show/segment only"]


def slugify_domain(name: str) -> str:
    # "The 831" -> "the831"; "Highway 1 Insider" -> "highway1insider"
    return re.sub(r"[^a-z0-9]", "", name.lower())


def quote(name: str) -> str:
    return urllib.parse.quote_plus(name)


def exact_quote(name: str) -> str:
    return urllib.parse.quote_plus(f'"{name}"')


def domain_exists(domain: str, timeout: float = 2.0) -> bool:
    # DNS existence proxy. Not the same as registrar availability.
    # If False: domain may be available OR registered without DNS.
    old_timeout = socket.getdefaulttimeout()
    socket.setdefaulttimeout(timeout)
    try:
        socket.getaddrinfo(domain, None)
        return True
    except Exception:
        return False
    finally:
        socket.setdefaulttimeout(old_timeout)


def safe_get_json(url: str, timeout: int = 10) -> Optional[dict]:
    if requests is None:
        return None
    try:
        r = requests.get(url, timeout=timeout, headers={"User-Agent": "BrandDiscoveryEngine/1.0"})
        if r.status_code == 200:
            return r.json()
    except Exception:
        return None
    return None


def apple_podcast_search(name: str) -> Tuple[int, str]:
    url = f"https://itunes.apple.com/search?term={quote(name)}&media=podcast&entity=podcast&limit=10"
    data = safe_get_json(url)
    if not data:
        return 0, url
    count = int(data.get("resultCount", 0))
    return count, url


def youtube_search(name: str, api_key: Optional[str]) -> Tuple[Optional[int], str]:
    search_url = f"https://www.youtube.com/results?search_query={quote(name)}"
    if not api_key or requests is None:
        return None, search_url
    api = (
        "https://www.googleapis.com/youtube/v3/search?"
        f"part=snippet&type=channel,video&maxResults=10&q={quote(name)}&key={api_key}"
    )
    data = safe_get_json(api)
    if not data:
        return None, search_url
    return len(data.get("items", [])), search_url


def serpapi_google_search(name: str, api_key: Optional[str]) -> Tuple[Optional[int], str]:
    google_url = f"https://www.google.com/search?q={exact_quote(name)}"
    if not api_key or requests is None:
        return None, google_url
    api = f"https://serpapi.com/search.json?engine=google&q={exact_quote(name)}&api_key={api_key}"
    data = safe_get_json(api)
    if not data:
        return None, google_url
    organic = data.get("organic_results", [])
    return len(organic), google_url


def build_links(name: str) -> Dict[str, str]:
    q = quote(name)
    q_exact = exact_quote(name)
    domain_base = slugify_domain(name)
    handle = re.sub(r"[^a-z0-9_]", "", name.lower().replace(" ", ""))

    return {
        "Google exact": f"https://www.google.com/search?q={q_exact}",
        "Google broad": f"https://www.google.com/search?q={q}",
        "Google images": f"https://www.google.com/search?tbm=isch&q={q}",
        "Namecheap": f"https://www.namecheap.com/domains/registration/results/?domain={domain_base}.com",
        "GoDaddy": f"https://www.godaddy.com/domainsearch/find?domainToCheck={domain_base}.com",
        "Porkbun": f"https://porkbun.com/checkout/search?q={domain_base}.com",
        "USPTO": f"https://tmsearch.uspto.gov/search/search-results?query={q}",
        "CA SOS": f"https://bizfileonline.sos.ca.gov/search/business",
        "YouTube": f"https://www.youtube.com/results?search_query={q}",
        "Instagram": f"https://www.instagram.com/{handle}/",
        "Facebook": f"https://www.facebook.com/search/top?q={q}",
        "TikTok": f"https://www.tiktok.com/search?q={q}",
        "LinkedIn": f"https://www.linkedin.com/search/results/all/?keywords={q}",
        "X": f"https://x.com/search?q={q}&src=typed_query",
        "Apple Podcasts": f"https://podcasts.apple.com/us/search?term={q}",
        "Spotify": f"https://open.spotify.com/search/{urllib.parse.quote(name)}",
    }


def score_row(row: dict) -> int:
    score = 50

    # Domain heuristic: if .com DNS resolves, treat as possible/taken and reduce.
    if row.get("com_dns_exists") == "Yes":
        score -= 15
    elif row.get("com_dns_exists") == "No":
        score += 8

    # Public API conflicts
    try:
        apple = int(row.get("apple_podcast_results") or 0)
        if apple >= 3:
            score -= 8
        elif apple == 0:
            score += 4
    except Exception:
        pass

    yt = row.get("youtube_results")
    if yt not in ("", None, "Manual"):
        try:
            yt_i = int(yt)
            if yt_i >= 5:
                score -= 5
            elif yt_i == 0:
                score += 3
        except Exception:
            pass

    google = row.get("google_exact_results")
    if google not in ("", None, "Manual"):
        try:
            g = int(google)
            if g >= 8:
                score -= 8
            elif g == 0:
                score += 5
        except Exception:
            pass

    # Brand qualities, simple defaults
    name = row.get("name", "").lower()
    if "highway 1" in name or "highway one" in name:
        score += 8
    if "central coast" in name:
        score += 7
    if "831" in name:
        score += 5
    if len(row.get("name","")) <= 20:
        score += 3
    if any(word in name for word in ["insider", "local", "journal", "weekend", "discoveries", "picks"]):
        score += 5
    if any(word in name for word in ["compass"]):
        score -= 8  # local real-estate concern based on Paul's observation

    return max(0, min(100, score))


def recommendation(score: int) -> str:
    if score >= 82:
        return "Strong candidate"
    if score >= 70:
        return "Possible"
    if score >= 55:
        return "Caution"
    return "Eliminate"


def research_name(name: str, youtube_key: Optional[str], serpapi_key: Optional[str]) -> dict:
    domain_base = slugify_domain(name)
    domains = {tld: f"{domain_base}.{tld}" for tld in TLDs}

    row = {
        "name": name,
        "domain_base": domain_base,
    }

    # DNS checks: safe and fast, but not definitive registrar availability.
    for tld, domain in domains.items():
        row[f"{tld}_domain"] = domain
        row[f"{tld}_dns_exists"] = "Yes" if domain_exists(domain) else "No"

    apple_count, apple_url = apple_podcast_search(name)
    row["apple_podcast_results"] = apple_count
    row["apple_podcast_url"] = apple_url

    yt_count, yt_url = youtube_search(name, youtube_key)
    row["youtube_results"] = yt_count if yt_count is not None else "Manual"
    row["youtube_url"] = yt_url

    google_count, google_url = serpapi_google_search(name, serpapi_key)
    row["google_exact_results"] = google_count if google_count is not None else "Manual"
    row["google_exact_url"] = google_url

    links = build_links(name)
    for k, v in links.items():
        row[f"link_{k}"] = v

    row["initial_score"] = score_row(row)
    row["recommendation"] = recommendation(row["initial_score"])
    row["manual_domain_status"] = "Unknown"
    row["manual_social_status"] = "Unknown"
    row["manual_trademark_risk"] = "Unknown"
    row["notes"] = ""

    return row


def read_candidates(path: Path) -> List[str]:
    text = path.read_text(encoding="utf-8-sig").splitlines()
    if not text:
        return []
    if "," in text[0].lower() or text[0].strip().lower() in ("name", "candidate"):
        rows = list(csv.DictReader(text))
        names = [r.get("name") or r.get("candidate") or r.get("Name") or "" for r in rows]
    else:
        names = text
    return [n.strip() for n in names if n.strip()]


def write_csv(rows: List[dict], path: Path):
    if not rows:
        return
    fields = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(rows: List[dict], path: Path):
    if Workbook is None:
        write_csv(rows, path.with_suffix(".csv"))
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Brand Research"

    # Keep key columns first and links grouped after
    preferred = [
        "name", "initial_score", "recommendation",
        "manual_domain_status", "manual_social_status", "manual_trademark_risk", "notes",
        "domain_base", "com_domain", "com_dns_exists", "co_domain", "co_dns_exists",
        "apple_podcast_results", "youtube_results", "google_exact_results",
        "link_Google exact", "link_Namecheap", "link_GoDaddy", "link_Porkbun",
        "link_USPTO", "link_YouTube", "link_Instagram", "link_Facebook",
        "link_TikTok", "link_LinkedIn", "link_X", "link_Apple Podcasts", "link_Spotify"
    ]
    all_keys = []
    for k in preferred:
        if k in rows[0]:
            all_keys.append(k)
    for k in rows[0].keys():
        if k not in all_keys:
            all_keys.append(k)

    ws.append(all_keys)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(k, "") for k in all_keys])

    # Hyperlink formatting
    for col_idx, key in enumerate(all_keys, 1):
        if key.startswith("link_"):
            for r in range(2, len(rows)+2):
                cell = ws.cell(r, col_idx)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    # Dropdowns
    def add_dropdown(col_name, options):
        if col_name not in all_keys:
            return
        col = all_keys.index(col_name) + 1
        col_letter = get_column_letter(col)
        dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}500")

    add_dropdown("manual_domain_status", MANUAL_STATUSES)
    add_dropdown("manual_social_status", MANUAL_STATUSES)
    add_dropdown("manual_trademark_risk", RISK_STATUSES)
    add_dropdown("recommendation", RECOMMENDATIONS)

    # Dashboard
    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "Brand Discovery Engine Dashboard"
    dash["A1"].font = Font(size=16, bold=True)
    dash["A3"] = "Sort the Brand Research tab by initial_score, then manually review the strongest candidates."
    dash["A5"] = "Important"
    dash["A5"].font = Font(bold=True)
    dash["A6"] = "DNS checks are not definitive domain availability. Confirm at a registrar before buying."
    dash["A7"] = "Social platforms and trademarks require manual review."
    dash["A9"] = "Best current candidates"
    dash["A9"].font = Font(bold=True)

    top = sorted(rows, key=lambda r: r.get("initial_score", 0), reverse=True)[:15]
    dash.append([])
    dash.append(["Rank", "Name", "Score", "Recommendation", ".com DNS", "Apple Podcast Results"])
    for i, r in enumerate(top, 1):
        dash.append([i, r["name"], r["initial_score"], r["recommendation"], r["com_dns_exists"], r["apple_podcast_results"]])

    for wsx in [ws, dash]:
        wsx.freeze_panes = "A2"
        for col in wsx.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[:100]:
                max_len = max(max_len, len(str(cell.value or "")))
            wsx.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 38)

    wb.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="candidates.csv")
    parser.add_argument("--output", default="brand_results.xlsx")
    parser.add_argument("--sleep", type=float, default=0.25, help="Pause between names")
    args = parser.parse_args()

    input_path = Path(args.input)
    names = read_candidates(input_path)

    youtube_key = os.getenv("YOUTUBE_API_KEY")
    serpapi_key = os.getenv("SERPAPI_KEY")

    rows = []
    for name in names:
        print(f"Researching: {name}")
        rows.append(research_name(name, youtube_key, serpapi_key))
        time.sleep(args.sleep)

    out = Path(args.output)
    if out.suffix.lower() == ".csv":
        write_csv(rows, out)
    else:
        write_xlsx(rows, out)

    print(f"Saved: {out.resolve()}")


if __name__ == "__main__":
    main()
